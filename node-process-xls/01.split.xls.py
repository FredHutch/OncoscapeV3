
import xlsx2csv
from openpyxl import load_workbook
import sys
import os

# file = 'D.xlsx'
file = sys.argv[1]

wb = load_workbook(filename=file, read_only=True)
sheetNames = wb.sheetnames
# converter = xlsx2csv.Xlsx2csv(file, sheetid=0, outputencoding='utf-8')
converter = xlsx2csv.Xlsx2csv(file, sheetid=0)
del wb


if not os.path.exists('output'):
    os.makedirs('output')

for idx, sheetName in enumerate(sheetNames):
    converter.convert('data', sheetid=idx)
    # converter.convert(sheetName + '.csv', sheetid=idx)
